"""
Core diff logic for the Excel Compare tool.

Design (per spec):
- Comparison is cell-by-cell values, same structure assumed.
- Rows are matched between the two files by a user-chosen key column
  (e.g. ID, Email) rather than row position — so reordered rows still
  match correctly.
- Output is a single downloadable .xlsx with differences highlighted:
    - "Summary"      — counts of changed cells, added rows, removed rows
    - "Changes"      — one row per changed cell: Key | Column | Old | New
    - "Added Rows"   — full rows present in File B but not File A
    - "Removed Rows" — full rows present in File A but not File B
    - "Highlighted"  — File B's full data, with changed cells shaded
                        yellow and added rows shaded green
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

YELLOW = PatternFill(start_color="FFF3C4", end_color="FFF3C4", fill_type="solid")
GREEN = PatternFill(start_color="D9F2D9", end_color="D9F2D9", fill_type="solid")
RED = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
HEADER_FILL = PatternFill(start_color="2E2E2E", end_color="2E2E2E", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def get_sheets(excel_path: str) -> list[str]:
    return pd.ExcelFile(excel_path).sheet_names


def get_columns(excel_path: str, sheet_name) -> list[str]:
    return list(pd.read_excel(excel_path, sheet_name=sheet_name, nrows=0).columns)


def _normalize(value):
    """Treat NaN/None as equivalent, and compare stripped string forms so
    trivial formatting differences (e.g. 42000 vs 42000.0) don't create
    false positives."""
    if pd.isna(value):
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def compare(file_a: str, sheet_a, file_b: str, sheet_b, key_column: str) -> dict:
    """
    Returns a dict with:
      columns: list of columns compared (common to both, excluding key)
      changes: list of {key, column, old, new}
      added_rows: list of dict (rows only in B)
      removed_rows: list of dict (rows only in A)
      df_b: the DataFrame for file B (used later to build the highlighted sheet)
    """
    df_a = pd.read_excel(file_a, sheet_name=sheet_a)
    df_b = pd.read_excel(file_b, sheet_name=sheet_b)

    if key_column not in df_a.columns:
        raise ValueError(f"Key column '{key_column}' not found in File A")
    if key_column not in df_b.columns:
        raise ValueError(f"Key column '{key_column}' not found in File B")

    common_columns = [c for c in df_a.columns if c in df_b.columns and c != key_column]
    if not common_columns:
        raise ValueError("No comparable columns in common between the two files (besides the key)")

    # Index by key column (as string, to avoid int/float/str key mismatches)
    df_a = df_a.copy()
    df_b = df_b.copy()
    df_a["_key"] = df_a[key_column].map(_normalize)
    df_b["_key"] = df_b[key_column].map(_normalize)

    a_by_key = {row["_key"]: row for _, row in df_a.iterrows()}
    b_by_key = {row["_key"]: row for _, row in df_b.iterrows()}

    a_keys = set(a_by_key.keys())
    b_keys = set(b_by_key.keys())

    common_keys = a_keys & b_keys
    removed_keys = a_keys - b_keys   # in A, not in B
    added_keys = b_keys - a_keys     # in B, not in A

    changes = []
    for key in common_keys:
        row_a = a_by_key[key]
        row_b = b_by_key[key]
        for col in common_columns:
            old_val = row_a[col]
            new_val = row_b[col]
            if _normalize(old_val) != _normalize(new_val):
                changes.append({
                    "key": row_b[key_column],
                    "column": col,
                    "old": "" if pd.isna(old_val) else old_val,
                    "new": "" if pd.isna(new_val) else new_val,
                })

    added_rows = [b_by_key[k].drop("_key").to_dict() for k in added_keys]
    removed_rows = [a_by_key[k].drop("_key").to_dict() for k in removed_keys]

    # Sort changes by key then column for a stable, readable report
    changes.sort(key=lambda c: (str(c["key"]), c["column"]))

    return {
        "key_column": key_column,
        "columns": common_columns,
        "changes": changes,
        "added_rows": added_rows,
        "removed_rows": removed_rows,
        "added_keys": added_keys,
        "df_b": df_b.drop(columns=["_key"]),
        "b_by_key": {k: v.drop("_key") for k, v in b_by_key.items()},
    }


def _autosize(ws):
    for col_cells in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max(length + 2, 10), 60)


def _write_header(ws, headers, row=1):
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")


def build_diff_report(file_a: str, sheet_a, file_b: str, sheet_b, key_column: str, output_path: str) -> str:
    result = compare(file_a, sheet_a, file_b, sheet_b, key_column)

    wb = Workbook()

    # --- Summary ---
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Excel Comparison Summary"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["Key column", result["key_column"]])
    ws.append(["Changed cells", len(result["changes"])])
    ws.append(["Rows added (in File B, not File A)", len(result["added_rows"])])
    ws.append(["Rows removed (in File A, not File B)", len(result["removed_rows"])])
    _autosize(ws)

    # --- Changes ---
    ws = wb.create_sheet("Changes")
    _write_header(ws, ["Key", "Column", "Old Value", "New Value"])
    for i, c in enumerate(result["changes"], start=2):
        ws.cell(row=i, column=1, value=c["key"])
        ws.cell(row=i, column=2, value=c["column"])
        ws.cell(row=i, column=3, value=c["old"]).fill = RED
        ws.cell(row=i, column=4, value=c["new"]).fill = GREEN
    _autosize(ws)

    # --- Added Rows ---
    ws = wb.create_sheet("Added Rows")
    if result["added_rows"]:
        headers = list(result["added_rows"][0].keys())
        _write_header(ws, headers)
        for i, row in enumerate(result["added_rows"], start=2):
            for j, h in enumerate(headers, start=1):
                ws.cell(row=i, column=j, value=row[h]).fill = GREEN
    else:
        ws.append(["No rows were added."])
    _autosize(ws)

    # --- Removed Rows ---
    ws = wb.create_sheet("Removed Rows")
    if result["removed_rows"]:
        headers = list(result["removed_rows"][0].keys())
        _write_header(ws, headers)
        for i, row in enumerate(result["removed_rows"], start=2):
            for j, h in enumerate(headers, start=1):
                ws.cell(row=i, column=j, value=row[h]).fill = RED
    else:
        ws.append(["No rows were removed."])
    _autosize(ws)

    # --- Highlighted (File B with changes/additions marked in place) ---
    ws = wb.create_sheet("Highlighted (File B)")
    df_b = result["df_b"]
    headers = list(df_b.columns)
    _write_header(ws, headers)

    # Build a quick lookup: (key, column) -> changed?
    changed_lookup = {(str(c["key"]), c["column"]) for c in result["changes"]}
    added_key_strs = {str(k) for k in result["added_keys"]}

    for i, (_, row) in enumerate(df_b.iterrows(), start=2):
        row_key = str(row[key_column])
        is_added_row = row_key in added_key_strs
        for j, col in enumerate(headers, start=1):
            cell = ws.cell(row=i, column=j, value=row[col] if not pd.isna(row[col]) else "")
            if is_added_row:
                cell.fill = GREEN
            elif (row_key, col) in changed_lookup:
                cell.fill = YELLOW
    _autosize(ws)

    wb.save(output_path)
    return output_path
